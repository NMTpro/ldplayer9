import logging
import os
import subprocess
import sys
import time
from datetime import datetime
import base64
import json
import logging


logging.basicConfig(level=logging.DEBUG)
try:
    with open(f'path.txt', 'r', encoding='utf8') as f:
        path_ld = f.read()
except:
    path_ld = ''
try:
    import pytz
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, protection
    import requests
    from datetime import datetime
    from PyQt5 import QtCore, QtWidgets
    from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QComboBox, QLabel
    from PyQt5.QtCore import pyqtSignal, QTimer, QThread, QObject
    import pyperclip
    import warnings
    import requests
    import numpy as np
    import cv2
    from PIL import Image
except:
    os.system("pip install openpyxl")
    os.system("pip install pyperclip")
    os.system("pip install PyQt5")
    os.system("pip install requests")
    os.system("pip install cv2")
    os.system("pip install numpy")
    os.system("pip install opencv-python")
    os.system('pip install Pillow')
    os.system("pip install pytz")
from PyQt5.QtGui import QIntValidator
import pytz
from datetime import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, protection
import requests
from datetime import datetime
import pyperclip
import numpy as np
import cv2
import requests
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtCore import pyqtSignal, QTimer

from guild import Ui_MainWindow

path_apk = './apk/'

with open('path.txt', 'r', encoding='utf8') as ffff:
    path = ffff.read()


def count_lines(file_path):
    with open(f'{file_path}.txt', 'r', encoding='utf8') as file:
        return sum(1 for line in file)


def open_ldplayer(ldplayer_path, index=2, minimize=True):
    ldplayer_exe = r'\dnplayer.exe'
    ldplayer_full_path = ldplayer_path + ldplayer_exe
    command_line = f'"{ldplayer_full_path}" index={index} {"minimize" if minimize else ""}'

    try:
        subprocess.Popen(command_line, shell=True)
    except Exception as e:
        print(f"Lỗi khi mở tab LDPlayer: {e}")


def close_ldplayer(ldplayer_path, index):
    try:
        ldplayer_processes = []
        for process in psutil.process_iter(['pid', 'name', 'cmdline']):
            if process.info['name'] == 'dnplayer.exe' and ldplayer_path in ' '.join(process.info['cmdline']):
                ldplayer_processes.append(process)

        if len(ldplayer_processes) >= index + 1:
            ldplayer_processes[index].terminate()
            print(f"Đã đóng tab LDPlayer có index={index}.")
        else:
            print(f"Không tìm thấy tab LDPlayer có index={index}.")
    except Exception as e:
        print(f"Lỗi khi đóng tab LDPlayer: {e}")


def Mo_Khoa(device):
    os.system(f'adb -s {device} shell input keyevent 4')
    time.sleep(2)
    os.system(f'adb -s {device} shell input swipe 270 819 270 48 2000')
    time.sleep(2)
    os.system(f'adb -s {device} shell input text "0000"')
    time.sleep(1)
    os.system(f'adb -s {device} shell input keyevent 66')
    time.sleep(1)


def Get_device():
    device = subprocess.check_output('adb devices')
    p = str(device).replace("b'List of devices attached", "").replace('\\r\\n', '').replace(' ', '')
    if len(p) > 0:
        listDevice = p.split('\\tdevice')
        listDevice.pop()
        return listDevice
    else:
        return []


class MainWindow1(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.check = 0
        self.button_d = QtWidgets.QPushButton()
        self.load_table()
        self.Them_TK()
        self.row = 0
        self.list_LD = []
        self.devices = []
        self.acc = {}
        self.ui.frame_2.setMaximumSize(QtCore.QSize(0, 10000000))
        self.ui.tableWidget_3.setMaximumWidth(0)
        self.ui.tableWidget_2.setColumnWidth(0, 120)

    def Them_TK(self):
        self.ui.tableWidget_4.setRowCount(4)
        self.lineTK = QtWidgets.QLineEdit()
        self.ui.tableWidget_4.setCellWidget(0, 0, self.lineTK)
        self.lineTK.setPlaceholderText('Nhập Tài Khoản')
        self.lineMK = QtWidgets.QLineEdit()
        self.ui.tableWidget_4.setCellWidget(1, 0, self.lineMK)
        self.lineMK.setPlaceholderText('Nhập Mật Khẩu')
        combo_box = QComboBox(self)
        combo_box.addItems(["Chọn Game", "G88", "M88", "R88", "W88", "V88", "G365", "M365", "R365", "W365"])
        self.ui.tableWidget_4.setCellWidget(2, 0, combo_box)
        self.button_OK = QtWidgets.QPushButton()
        self.button_OK.setText('Thêm')
        self.ui.tableWidget_4.setCellWidget(3, 0, self.button_OK)
        self.button_OK.clicked.connect(self.button_ok_clicked)

    def button_ok_clicked(self):
        tai_khoan = self.lineTK.text()
        mat_khau = self.lineMK.text()
        game_duoc_chon = self.ui.tableWidget_4.cellWidget(2, 0).currentText()
        if tai_khoan != '' and mat_khau != '' and game_duoc_chon != 'Chọn Game' and len(self.devices) != 0:
            self.buttonz1.setText(f'Tài Khoản: {tai_khoan.strip()} Game: {game_duoc_chon} Đã Được Thêm Thành Công !')
            self.lineTK.setText('')
            self.lineMK.setText('')
            self.ui.tableWidget_4.cellWidget(2, 0).setCurrentIndex(0)
            min_lines_file = min(self.devices, key=count_lines)
            min_lines_file_path = min_lines_file
            with open(f'{min_lines_file_path}.txt', 'a', encoding='utf8') as file:
                new_line = f'{tai_khoan.strip()}|{mat_khau.strip()}|{game_duoc_chon}'
                file.write(new_line + '\n')
            with open(f'{min_lines_file_path}.txt', 'r', encoding='utf8') as f:
                addd = f.read().splitlines()
        else:
            self.buttonz1.setText('Không Hợp Lệ')

    def load_table(self):
        for ld in range(3):
            row = self.ui.tableWidget.rowCount()
            self.ui.tableWidget.insertRow(row)
            self.button = QtWidgets.QPushButton()
            self.button.setText('Mở LDPlayer')
            self.ui.tableWidget.setCellWidget(1, 0, self.button)
            self.button.clicked.connect(self.start_select)
            self.button1 = QtWidgets.QPushButton()
            self.button1.setText('Get DEVICE')
            self.button1.setEnabled(False)
            self.ui.tableWidget.setCellWidget(2, 0, self.button1)
            self.button1.clicked.connect(self.get_device)
            self.line = QtWidgets.QLineEdit()
            validator = QIntValidator()
            self.line.setValidator(validator)
            self.ui.tableWidget.setCellWidget(0, 0, self.line)
            self.line.setPlaceholderText('số luồng')
            self.buttonz = QtWidgets.QPushButton()
            self.buttonz.setText('')
            self.ui.tableWidget.setCellWidget(0, 2, self.buttonz)
            self.buttonz1 = QtWidgets.QPushButton()
            self.buttonz1.setText('')
            self.ui.tableWidget.setCellWidget(1, 2, self.buttonz1)
            self.bnt_menu2 = QtWidgets.QPushButton()
            self.bnt_menu2.setText('Mở Menu device')
            self.ui.tableWidget.setCellWidget(0, 1, self.bnt_menu2)
            self.bnt_menu2.clicked.connect(self.Menu_device)
            self.bnt_menu3 = QtWidgets.QPushButton()
            self.bnt_menu3.setText('Menu Thêm TK')
            self.ui.tableWidget.setCellWidget(1, 1, self.bnt_menu3)
            self.bnt_menu3.clicked.connect(self.Menu_add_tk)
        tai_khoan = QtWidgets.QPushButton()
        tai_khoan.setText('Trống')
        self.ui.tableWidget_3.setCellWidget(0, 0, tai_khoan)
        self.trang_thai1 = QtWidgets.QPushButton()
        self.ui.tableWidget_3.setCellWidget(0, 2, self.trang_thai1)
        game = QtWidgets.QPushButton()
        game.setText('Trống')
        self.ui.tableWidget_3.setCellWidget(0, 1, game)
        self.button_d1 = QtWidgets.QPushButton()
        self.button_d1.setText('Trở Về')
        self.ui.tableWidget_3.setCellWidget(0, 3, self.button_d1)
        self.button_d1.clicked.connect(self.back)

    def click_bnt_device(self):
        sender_button = self.sender()
        if isinstance(sender_button, QtWidgets.QPushButton):
            button_text = sender_button.text()
            try:
                with open(f'{button_text}.txt', 'r', encoding='utf8') as f:
                    tk = f.readlines()
            except:
                tk = []

            self.ui.tableWidget_3.setMaximumWidth(100000)
            self.ui.tableWidget_2.setMaximumWidth(0)

            self.ui.tableWidget_3.setRowCount(1)
            self.button_d1 = QtWidgets.QPushButton()
            self.button_d1.setText('Trở Về')
            self.ui.tableWidget_3.setCellWidget(0, 3, self.button_d1)
            self.button_d1.clicked.connect(self.back)

            if len(tk) > 0:
                self.ui.tableWidget_3.setRowCount(0)
                try:
                    with open(f'{button_text}.txt', 'r', encoding='utf8') as file:
                        for line in file:
                            elements = line.strip().split('|')
                            row = self.ui.tableWidget_3.rowCount()
                            self.ui.tableWidget_3.insertRow(row)
                            tai_khoan = QtWidgets.QPushButton()
                            tai_khoan.setText(elements[0])
                            self.ui.tableWidget_3.setCellWidget(row, 0, tai_khoan)
                            game = QtWidgets.QPushButton()
                            game.setText(elements[2])
                            self.ui.tableWidget_3.setCellWidget(row, 1, game)
                            self.trang_thai1 = QtWidgets.QPushButton()
                            self.trang_thai1.setText(str(self.acc[elements[0]]))
                            self.ui.tableWidget_3.setCellWidget(row, 2, self.trang_thai1)
                            self.button_d1 = QtWidgets.QPushButton()
                            self.button_d1.setText('Trở Về')
                            self.ui.tableWidget_3.setCellWidget(0, 3, self.button_d1)
                            self.button_d1.clicked.connect(self.back)
                except Exception as e:
                    self.button_d1 = QtWidgets.QPushButton()
                    self.button_d1.setText('Trở Về')
                    self.ui.tableWidget_3.setCellWidget(0, 3, self.button_d1)
                    self.button_d1.clicked.connect(self.back)
                    print(f"Error: {e}")
            else:
                tai_khoan = QtWidgets.QPushButton()
                tai_khoan.setText('Trống')
                self.ui.tableWidget_3.setCellWidget(0, 0, tai_khoan)
                self.trang_thai1 = QtWidgets.QPushButton()
                self.trang_thai1.setText('Trống')
                self.ui.tableWidget_3.setCellWidget(0, 2, self.trang_thai1)
                game = QtWidgets.QPushButton()
                game.setText('Trống')
                self.ui.tableWidget_3.setCellWidget(0, 1, game)
                self.button_d1 = QtWidgets.QPushButton()
                self.button_d1.setText('Trở Về')
                self.ui.tableWidget_3.setCellWidget(0, 3, self.button_d1)
                self.button_d1.clicked.connect(self.back)

    def thong_tin_acc(self, msg):
        self.trang_thai1.setText(str(msg[1]))
        self.acc[msg[0]] = msg[2]

    def device(self, msg):
        self.ui.tableWidget_2.setRowCount(0)
        self.devices = msg
        for i in msg:
            row = self.ui.tableWidget_2.rowCount()
            self.ui.tableWidget_2.insertRow(row)
            self.button_d = QtWidgets.QPushButton()
            self.button_d.setText(str(i))
            self.ui.tableWidget_2.setCellWidget(row, 1, self.button_d)
            self.button_d.clicked.connect(self.click_bnt_device)

    def Menu_device(self):
        if self.bnt_menu2.text() == 'Mở Menu device':
            self.ui.tableWidget_2.setRowCount(0)
            for i in self.devices:
                row = self.ui.tableWidget_2.rowCount()
                self.ui.tableWidget_2.insertRow(row)
                self.button_d = QtWidgets.QPushButton()
                self.button_d.setText(str(i))
                self.ui.tableWidget_2.setCellWidget(row, 1, self.button_d)
                self.button_d.clicked.connect(self.click_bnt_device)
                try:
                    with open(f'{i}.txt', 'r', encoding='utf8') as f:
                        tk = f.readlines()
                except:
                    tk = []
                label = QLabel(f"{len(tk)}")
                self.ui.tableWidget_2.setCellWidget(row, 0, label)
            self.ui.frame_2.setMaximumSize(100000000, 10000000)
            self.ui.tableWidget_4.setMaximumWidth(0)
            self.ui.tableWidget_2.setMaximumWidth(1000000)
            self.bnt_menu2.setText('Ẩn Menu device')
            self.bnt_menu3.setText('Menu Thêm TK')
        else:
            self.ui.frame_2.setMaximumSize(QtCore.QSize(0, 10000000))
            self.ui.tableWidget_4.setMaximumWidth(0)
            self.ui.tableWidget_3.setMaximumWidth(0)
            self.bnt_menu2.setText('Mở Menu device')

    def Menu_add_tk(self):
        if self.bnt_menu3.text() == 'Menu Thêm TK':
            self.ui.frame_2.setMaximumSize(100000000, 10000000)
            self.ui.tableWidget_4.setMaximumWidth(1000000)
            self.ui.tableWidget_2.setMaximumWidth(0)
            self.ui.tableWidget_3.setMaximumWidth(0)
            self.bnt_menu3.setText('Ẩn Menu Thêm TK')
            self.bnt_menu2.setText('Mở Menu device')
        else:
            self.ui.frame_2.setMaximumSize(QtCore.QSize(0, 10000000))
            self.ui.tableWidget_4.setMaximumWidth(0)
            self.ui.tableWidget_2.setMaximumWidth(1000000)
            self.bnt_menu3.setText('Menu Thêm TK')

    def back(self):
        if self.button_d1.text() == 'Trở Về':
            self.ui.tableWidget_3.setMaximumWidth(0)
            self.ui.tableWidget_2.setMaximumWidth(100000)

    def start_select(self):
        if self.line.text() != '':
            self.button.setEnabled(False)
            self.b = self.check
            self.a = self.check + int(self.line.text())
            while self.check < self.a:
                open_ldplayer(path, index=self.check, minimize=True)
                self.check +=1
            self.button1.setEnabled(True)

    def get_device(self):
        self.button.setEnabled(False)
        self.button1.setEnabled(False)
        if int(self.line.text()) > 0:
            self.thread = Get_Device(mainWindow=self)
            self.thread.signal.connect(self.thong_bao)
            self.thread.signal1.connect(self.device)
            self.thread.start()

    def thong_bao(self, msg):
        self.buttonz.setText(msg)

    def get_tk(self, device):
        try:
            with open(f'{device}.txt', 'r', encoding='utf-8') as f:
                req = f.read().splitlines()
        except:
            req = ""
        username = []
        passw = []
        gamea = []
        for i in range(len(req)):
            tk = req[i].split('|')[0]
            mk = req[i].split('|')[1]
            game = req[i].split('|')[2]
            username.append(tk)
            passw.append(mk)
            gamea.append(game)
        return username, passw, gamea


class Get_Device(QThread):
    signal = pyqtSignal(str)
    signal1 = pyqtSignal(list)

    def __init__(self, parent=None, mainWindow=None, index=0):
        super().__init__(parent)
        self.mainWindow = mainWindow
        self.index = index
        self.device = []
        self.count_device = 0

    def run(self):
        if int(self.mainWindow.line.text()) > 0:
            self.signal.emit('load giả lập')
            if self.mainWindow:
                while True:
                    self.count_device += len(self.device)
                    self.device = Get_device()
                    a = self.device
                    if len(self.device) == self.mainWindow.a:
                        self.signal.emit(f'Load Xong {len(self.device)} device')
                        self.signal1.emit(self.device)
                        break
                    self.signal.emit(f'Load được {len(self.device)} device')
                    self.signal1.emit(self.device)
                if len(self.device) > 0:
                    while self.mainWindow.b < len(self.device):
                        Mo_Khoa(self.device[self.mainWindow.b])
                        with open(f'{self.device[self.mainWindow.b]}.txt', 'w', encoding='utf8') as lz:
                            lz.write('')
                        self.signal.emit(f'Mở Khóa {self.device[self.mainWindow.b]} thành công !')
                        self.mainWindow.b += 1
            self.signal.emit(f'Cấu Hình Xong {len(self.device)} Device')
            self.signal1.emit(self.device)
            self.mainWindow.button.setEnabled(True)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_win1 = MainWindow1()
    main_win1.show()
    sys.exit(app.exec())
